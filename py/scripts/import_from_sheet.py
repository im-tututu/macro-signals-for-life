"""Import historical data from Google Sheets into SQLite."""

#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import sqlite3
import sys
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

import openpyxl

CURRENT_DIR = Path(__file__).resolve().parent
PY_ROOT = CURRENT_DIR.parent
SRC_DIR = PY_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from core.config import AppConfig, load_local_env


@dataclass(frozen=True)
class SheetSpec:
    sheet_name: str
    staging_table: str
    header_aliases: dict[str, str]
    required: bool = True
    header_row: int | None = 1
    auto_find_header: bool = False
    min_header_hits: int = 2


def parse_args() -> argparse.Namespace:
    load_local_env()
    cfg = AppConfig.load()
    parser = argparse.ArgumentParser(description="Import historical Google Sheet data into SQLite staging/raw tables.")
    parser.add_argument(
        "--xlsx",
        default=str(cfg.workbook_path) if cfg.workbook_path else "宏观观察.xlsx",
        help="Path to workbook (.xlsx). Default: read WORKBOOK_PATH from .env, or 宏观观察.xlsx",
    )
    parser.add_argument(
        "--db",
        default=str(cfg.db_path),
        help="Path to SQLite database. Default: read DB_PATH from .env, or py/data/db.sqlite",
    )
    parser.add_argument(
        "--sheet-max-scan-rows",
        type=int,
        default=10,
        help="Max rows to scan when auto-detecting header row. Default: 10",
    )
    return parser.parse_args()


SHEET_SPECS: list[SheetSpec] = [
    SheetSpec(
        sheet_name="运行日志",
        staging_table="stg_sheet_run_log",
        header_aliases={
            "timestamp": "timestamp",
            "job_name": "job_name",
            "status": "status",
            "message": "message",
            "detail": "detail",
        },
    ),
    SheetSpec(
        sheet_name="原始_收益率曲线",
        staging_table="stg_sheet_raw_bond_curve",
        header_aliases={
            "date": "date",
            "curve": "curve",
            "y_0": "y_0",
            "y_0.08": "y_0_08",
            "y_0.17": "y_0_17",
            "y_0.25": "y_0_25",
            "y_0.5": "y_0_5",
            "y_0.75": "y_0_75",
            "y_1": "y_1",
            "y_2": "y_2",
            "y_3": "y_3",
            "y_4": "y_4",
            "y_5": "y_5",
            "y_6": "y_6",
            "y_7": "y_7",
            "y_8": "y_8",
            "y_9": "y_9",
            "y_10": "y_10",
            "y_15": "y_15",
            "y_20": "y_20",
            "y_30": "y_30",
            "y_40": "y_40",
            "y_50": "y_50",
        },
    ),
    SheetSpec(
        sheet_name="原始_海外宏观",
        staging_table="stg_sheet_raw_overseas_macro",
        header_aliases={
            "date": "date",
            "fed_upper": "fed_upper",
            "fed_lower": "fed_lower",
            "sofr": "sofr",
            "ust_2y": "ust_2y",
            "ust_10y": "ust_10y",
            "us_real_10y": "us_real_10y",
            "usd_broad": "usd_broad",
            "usd_cny": "usd_cny",
            "gold": "gold",
            "wti": "wti",
            "brent": "brent",
            "copper": "copper",
            "vix": "vix",
            "spx": "spx",
            "nasdaq_100": "nasdaq_100",
            "source": "source",
            "fetched_at": "fetched_at",
        },
    ),
    SheetSpec(
        sheet_name="原始_政策利率",
        staging_table="stg_sheet_raw_policy_rate",
        header_aliases={
            "date": "date",
            "type": "type",
            "term": "term",
            "rate": "rate",
            "amount": "amount",
            "source": "source",
            "fetched_at": "fetched_at",
            "note": "note",
        },
    ),
    SheetSpec(
        sheet_name="原始_资金面",
        staging_table="stg_sheet_raw_money_market",
        header_aliases={
            "date": "date",
            "showdatecn": "show_date_cn",
            "source_url": "source_url",
            "dr001_weightedrate": "dr001_weighted_rate",
            "dr001_latestrate": "dr001_latest_rate",
            "dr001_avgprd": "dr001_avg_prd",
            "dr007_weightedrate": "dr007_weighted_rate",
            "dr007_latestrate": "dr007_latest_rate",
            "dr007_avgprd": "dr007_avg_prd",
            "dr014_weightedrate": "dr014_weighted_rate",
            "dr014_latestrate": "dr014_latest_rate",
            "dr014_avgprd": "dr014_avg_prd",
            "dr021_weightedrate": "dr021_weighted_rate",
            "dr021_latestrate": "dr021_latest_rate",
            "dr021_avgprd": "dr021_avg_prd",
            "dr1m_weightedrate": "dr1m_weighted_rate",
            "dr1m_latestrate": "dr1m_latest_rate",
            "dr1m_avgprd": "dr1m_avg_prd",
            "fetched_at": "fetched_at",
        },
    ),
    SheetSpec(
        sheet_name="原始_货币",
        staging_table="stg_sheet_raw_money_legacy",
        header_aliases={
            "date": "date",
            "showdatecn": "show_date_cn",
            "source_url": "source_url",
            "dr001_weightedrate": "dr001_weighted_rate",
            "dr001_latestrate": "dr001_latest_rate",
            "dr001_avgprd": "dr001_avg_prd",
            "dr007_weightedrate": "dr007_weighted_rate",
            "dr007_latestrate": "dr007_latest_rate",
            "dr007_avgprd": "dr007_avg_prd",
            "dr014_weightedrate": "dr014_weighted_rate",
            "dr014_latestrate": "dr014_latest_rate",
            "dr014_avgprd": "dr014_avg_prd",
            "dr021_weightedrate": "dr021_weighted_rate",
            "dr021_latestrate": "dr021_latest_rate",
            "dr021_avgprd": "dr021_avg_prd",
            "dr1m_weightedrate": "dr1m_weighted_rate",
            "dr1m_latestrate": "dr1m_latest_rate",
            "dr1m_avgprd": "dr1m_avg_prd",
            "fetched_at": "fetched_at",
        },
        required=False,
    ),
    SheetSpec(
        sheet_name="原始_国债期货",
        staging_table="stg_sheet_raw_futures",
        header_aliases={
            "date": "date",
            "t0_last": "t0_last",
            "tf0_last": "tf0_last",
            "source": "source",
            "fetched_at": "fetched_at",
        },
    ),
    SheetSpec(
        sheet_name="原始_民生与资产价格",
        staging_table="stg_sheet_raw_life_asset",
        header_aliases={
            "date": "date",
            "mortgage_rate_est": "mortgage_rate_est",
            "house_price_tier1": "house_price_tier1",
            "house_price_tier2": "house_price_tier2",
            "house_price_nbs_70city": "house_price_nbs_70city",
            "gold_cny": "gold_cny",
            "money_fund_7d": "money_fund_7d",
            "deposit_1y": "deposit_1y",
            "source": "source",
            "fetched_at": "fetched_at",
        },
    ),
    SheetSpec(
        sheet_name="原始_指数ETF",
        staging_table="stg_sheet_raw_jisilu_etf",
        header_aliases={
            "snapshot_date": "snapshot_date",
            "fetched_at": "fetched_at",
            "fund_id": "fund_id",
            "fund_nm": "fund_nm",
            "index_nm": "index_nm",
            "issuer_nm": "issuer_nm",
            "price": "price",
            "increase_rt": "increase_rt",
            "volume_wan": "volume_wan",
            "amount_yi": "amount_yi",
            "unit_total_yi": "unit_total_yi",
            "discount_rt": "discount_rt",
            "fund_nav": "fund_nav",
            "nav_dt": "nav_dt",
            "estimate_value": "estimate_value",
            "creation_unit": "creation_unit",
            "pe": "pe",
            "pb": "pb",
            "last_time": "last_time",
            "last_est_time": "last_est_time",
            "is_qdii": "is_qdii",
            "is_t0": "is_t0",
            "apply_fee": "apply_fee",
            "redeem_fee": "redeem_fee",
            "records_total": "records_total",
            "source_url": "source_url",
        },
    ),
    SheetSpec(
        sheet_name="原始_债券指数特征",
        staging_table="stg_sheet_raw_bond_index",
        header_aliases={
            "trade_date": "trade_date",
            "index_name": "index_name",
            "index_code": "index_code",
            "provider": "provider",
            "type_lv1": "type_lv1",
            "type_lv2": "type_lv2",
            "type_lv3": "type_lv3",
            "source_url": "source_url",
            "data_date": "data_date",
            "dm": "dm",
            "y": "y",
            "cons_number": "cons_number",
            "d": "d",
            "v": "v",
            "fetch_status": "fetch_status",
            "raw_json": "raw_json",
            "fetched_at": "fetched_at",
            "error": "error",
        },
    ),
    SheetSpec(
        sheet_name="配置_债券指数清单",
        staging_table="stg_sheet_cfg_bond_index_list",
        header_aliases={
            "指数": "index_name",
            "代码": "index_code",
            "备注": "note",
            "类型": "type_lv1",
            "类型-二级": "type_lv2",
            "类型-三级": "type_lv3",
            "指数发行公司": "provider",
            "数据日期": "data_date",
            "dm（修正久期）": "dm",
            "y（到期收益率）": "y",
            "consnumber（成分债券数量）": "cons_number",
            "d（久期）": "d",
            "v（凸性）": "v",
        },
        header_row=None,
        auto_find_header=True,
        min_header_hits=4,
    ),
    SheetSpec(
        sheet_name="映射_Jisilu债券指数候选",
        staging_table="stg_sheet_map_jisilu_bond_index_candidate",
        header_aliases={
            "index_name": "index_name",
            "match_key": "match_key",
            "provider_guess": "provider_guess",
            "type_lv1": "type_lv1",
            "type_lv2": "type_lv2",
            "type_lv3": "type_lv3",
            "sample_fund_code": "sample_fund_code",
            "sample_fund_name": "sample_fund_name",
            "source_data_date": "source_data_date",
            "source_sheet": "source_sheet",
            "note": "note",
            "updated_at": "updated_at",
        },
    ),
]


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\n", "")
    text = text.replace("\r", "")
    text = text.replace(" ", "")
    return text.lower()


def to_text(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.replace(microsecond=0).isoformat(sep=" ")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, time):
        if value.microsecond:
            value = value.replace(microsecond=0)
        return value.isoformat()

    if isinstance(value, bool):
        return "1" if value else "0"

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if math.isnan(value):
            return None
        return format(value, ".15g")

    text = str(value).strip()
    return text or None


def quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def get_table_columns(conn: sqlite3.Connection, table_name: str) -> list[str]:
    rows = conn.execute(f"PRAGMA table_info({quote_ident(table_name)})").fetchall()
    if not rows:
        raise RuntimeError(f"Table not found: {table_name}")
    return [row[1] for row in rows]


def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, spec: SheetSpec, max_scan_rows: int) -> int:
    expected = {normalize_header(k) for k in spec.header_aliases}
    best_row = None
    best_hits = -1

    for row_idx in range(1, max_scan_rows + 1):
        values = [cell.value for cell in ws[row_idx]]
        hits = sum(1 for v in values if normalize_header(v) in expected)
        if hits > best_hits:
            best_hits = hits
            best_row = row_idx

    if best_row is None or best_hits < spec.min_header_hits:
        raise RuntimeError(
            f"Could not find header row for sheet={spec.sheet_name!r}; best_hits={best_hits}, required={spec.min_header_hits}"
        )
    return best_row


def build_header_map(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    spec: SheetSpec,
    header_row: int,
    table_columns: set[str],
) -> dict[int, str]:
    mapped: dict[int, str] = {}
    raw_headers = [cell.value for cell in ws[header_row]]

    seen_deposit_1y = 0
    seen_source = 0
    seen_fetched_at = 0

    for idx, raw_header in enumerate(raw_headers, start=1):
        key = normalize_header(raw_header)
        if not key:
            if spec.staging_table == "stg_sheet_raw_policy_rate":
                if "extra_1" in table_columns and "extra_1" not in mapped.values():
                    mapped[idx] = "extra_1"
                elif "extra_2" in table_columns and "extra_2" not in mapped.values():
                    mapped[idx] = "extra_2"
            continue

        target = spec.header_aliases.get(key)
        if target is None:
            continue

        if spec.staging_table == "stg_sheet_raw_life_asset":
            if target == "deposit_1y":
                seen_deposit_1y += 1
                target = "deposit_1y" if seen_deposit_1y == 1 else "deposit_1y_dup"
            elif target == "source":
                seen_source += 1
                target = "source" if seen_source == 1 else "source_dup"
            elif target == "fetched_at":
                seen_fetched_at += 1
                target = "fetched_at" if seen_fetched_at == 1 else "fetched_at_dup"

        if target in table_columns:
            mapped[idx] = target

    if len(mapped) < spec.min_header_hits:
        raise RuntimeError(
            f"Header mapping too weak for sheet={spec.sheet_name!r}; matched_columns={len(mapped)}"
        )

    return mapped


def is_effectively_empty(values: Iterable[Any]) -> bool:
    for v in values:
        if v is not None and str(v).strip() != "":
            return False
    return True


def clear_table(conn: sqlite3.Connection, table_name: str) -> None:
    conn.execute(f"DELETE FROM {quote_ident(table_name)}")


def insert_rows(conn: sqlite3.Connection, table_name: str, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0

    all_columns = list(rows[0].keys())
    placeholders = ", ".join(["?"] * len(all_columns))
    sql = f"""
        INSERT INTO {quote_ident(table_name)} ({", ".join(quote_ident(c) for c in all_columns)})
        VALUES ({placeholders})
    """
    params = [tuple(row.get(col) for col in all_columns) for row in rows]
    conn.executemany(sql, params)
    return len(rows)


def stage_sheet(
    conn: sqlite3.Connection,
    wb: openpyxl.Workbook,
    spec: SheetSpec,
    max_scan_rows: int,
) -> int:
    if spec.sheet_name not in wb.sheetnames:
        if spec.required:
            raise RuntimeError(f"Required sheet not found: {spec.sheet_name}")
        print(f"[SKIP] Missing optional sheet: {spec.sheet_name}")
        return 0

    ws = wb[spec.sheet_name]
    table_columns = get_table_columns(conn, spec.staging_table)
    table_column_set = set(table_columns)

    header_row = (
        find_header_row(ws, spec, max_scan_rows)
        if spec.auto_find_header
        else int(spec.header_row or 1)
    )

    header_map = build_header_map(ws, spec, header_row, table_column_set)

    clear_table(conn, spec.staging_table)

    staged_rows: list[dict[str, Any]] = []
    for row_idx, row_values in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if is_effectively_empty(row_values):
            continue

        payload: dict[str, Any] = {"source_row_num": row_idx}
        for col_idx, target_col in header_map.items():
            raw_value = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None
            payload[target_col] = to_text(raw_value)

        staged_rows.append(payload)

    inserted = insert_rows(conn, spec.staging_table, staged_rows)
    print(f"[STAGE] {spec.sheet_name} -> {spec.staging_table} | header_row={header_row} | rows={inserted}")
    return inserted


def cast_real(expr: str) -> str:
    return f"CAST(NULLIF(TRIM({expr}), '') AS REAL)"


def cast_int(expr: str) -> str:
    return f"CAST(NULLIF(TRIM({expr}), '') AS INTEGER)"


def nonblank(expr: str) -> str:
    return f"NULLIF(TRIM({expr}), '')"


def merge_table(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> int:
    conn.execute(sql, params)
    return int(conn.execute("SELECT changes()").fetchone()[0])


def merge_staged_data(conn: sqlite3.Connection) -> dict[str, int]:
    migrated_at = datetime.now().replace(microsecond=0).isoformat(sep=" ")
    counts: dict[str, int] = {}

    counts["run_log"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO run_log (
            timestamp, job_name, status, message, detail,
            source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('timestamp')},
            {nonblank('job_name')},
            {nonblank('status')},
            {nonblank('message')},
            {nonblank('detail')},
            ?, source_row_num, ?
        FROM stg_sheet_run_log
        WHERE {nonblank('timestamp')} IS NOT NULL
          AND {nonblank('job_name')} IS NOT NULL
        """,
        ("运行日志", migrated_at),
    )

    counts["cfg_bond_index_list"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO cfg_bond_index_list (
            index_name, index_code, note, type_lv1, type_lv2, type_lv3, provider,
            data_date, dm, y, cons_number, d, v,
            source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('index_name')},
            {nonblank('index_code')},
            {nonblank('note')},
            {nonblank('type_lv1')},
            {nonblank('type_lv2')},
            {nonblank('type_lv3')},
            {nonblank('provider')},
            {nonblank('data_date')},
            {cast_real('dm')},
            {cast_real('y')},
            {cast_int('cons_number')},
            {cast_real('d')},
            {cast_real('v')},
            ?, source_row_num, ?
        FROM stg_sheet_cfg_bond_index_list
        WHERE {nonblank('index_name')} IS NOT NULL
        """,
        ("配置_债券指数清单", migrated_at),
    )

    counts["map_jisilu_bond_index_candidate"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO map_jisilu_bond_index_candidate (
            index_name, match_key, provider_guess, type_lv1, type_lv2, type_lv3,
            sample_fund_code, sample_fund_name, source_data_date, source_sheet, note, updated_at
        )
        SELECT
            {nonblank('index_name')},
            {nonblank('match_key')},
            {nonblank('provider_guess')},
            {nonblank('type_lv1')},
            {nonblank('type_lv2')},
            {nonblank('type_lv3')},
            {nonblank('sample_fund_code')},
            {nonblank('sample_fund_name')},
            {nonblank('source_data_date')},
            {nonblank('source_sheet')},
            {nonblank('note')},
            COALESCE({nonblank('updated_at')}, ?)
        FROM stg_sheet_map_jisilu_bond_index_candidate
        WHERE {nonblank('index_name')} IS NOT NULL
        """,
        (migrated_at,),
    )

    counts["raw_bond_curve"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO raw_bond_curve (
            date, curve,
            y_0, y_0_08, y_0_17, y_0_25, y_0_5, y_0_75,
            y_1, y_2, y_3, y_4, y_5, y_6, y_7, y_8, y_9, y_10, y_15, y_20, y_30, y_40, y_50,
            source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('date')},
            {nonblank('curve')},
            {cast_real('y_0')}, {cast_real('y_0_08')}, {cast_real('y_0_17')}, {cast_real('y_0_25')},
            {cast_real('y_0_5')}, {cast_real('y_0_75')}, {cast_real('y_1')}, {cast_real('y_2')},
            {cast_real('y_3')}, {cast_real('y_4')}, {cast_real('y_5')}, {cast_real('y_6')},
            {cast_real('y_7')}, {cast_real('y_8')}, {cast_real('y_9')}, {cast_real('y_10')},
            {cast_real('y_15')}, {cast_real('y_20')}, {cast_real('y_30')}, {cast_real('y_40')}, {cast_real('y_50')},
            ?, source_row_num, ?
        FROM stg_sheet_raw_bond_curve
        WHERE {nonblank('date')} IS NOT NULL
          AND {nonblank('curve')} IS NOT NULL
        """,
        ("原始_收益率曲线", migrated_at),
    )

    counts["raw_overseas_macro"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO raw_overseas_macro (
            date, fed_upper, fed_lower, sofr, ust_2y, ust_10y, us_real_10y,
            usd_broad, usd_cny, gold, wti, brent, copper, vix, spx, nasdaq_100, source, fetched_at,
            source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('date')},
            {cast_real('fed_upper')}, {cast_real('fed_lower')}, {cast_real('sofr')},
            {cast_real('ust_2y')}, {cast_real('ust_10y')}, {cast_real('us_real_10y')},
            {cast_real('usd_broad')}, {cast_real('usd_cny')}, {cast_real('gold')},
            {cast_real('wti')}, {cast_real('brent')}, {cast_real('copper')},
            {cast_real('vix')}, {cast_real('spx')}, {cast_real('nasdaq_100')},
            {nonblank('source')}, {nonblank('fetched_at')},
            ?, source_row_num, ?
        FROM stg_sheet_raw_overseas_macro
        WHERE {nonblank('date')} IS NOT NULL
        """,
        ("原始_海外宏观", migrated_at),
    )

    counts["raw_policy_rate"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO raw_policy_rate (
            date, type, term, rate, amount, source, fetched_at, note,
            source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('date')},
            {nonblank('type')},
            {nonblank('term')},
            {cast_real('rate')},
            {cast_real('amount')},
            {nonblank('source')},
            {nonblank('fetched_at')},
            {nonblank('note')},
            ?, source_row_num, ?
        FROM stg_sheet_raw_policy_rate
        WHERE {nonblank('date')} IS NOT NULL
          AND {nonblank('type')} IS NOT NULL
          AND {nonblank('term')} IS NOT NULL
        """,
        ("原始_政策利率", migrated_at),
    )

    counts["raw_money_market_from_current"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO raw_money_market (
            date, show_date_cn, source_url,
            dr001_weighted_rate, dr001_latest_rate, dr001_avg_prd,
            dr007_weighted_rate, dr007_latest_rate, dr007_avg_prd,
            dr014_weighted_rate, dr014_latest_rate, dr014_avg_prd,
            dr021_weighted_rate, dr021_latest_rate, dr021_avg_prd,
            dr1m_weighted_rate, dr1m_latest_rate, dr1m_avg_prd,
            fetched_at, source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('date')}, {nonblank('show_date_cn')}, {nonblank('source_url')},
            {cast_real('dr001_weighted_rate')}, {cast_real('dr001_latest_rate')}, {cast_real('dr001_avg_prd')},
            {cast_real('dr007_weighted_rate')}, {cast_real('dr007_latest_rate')}, {cast_real('dr007_avg_prd')},
            {cast_real('dr014_weighted_rate')}, {cast_real('dr014_latest_rate')}, {cast_real('dr014_avg_prd')},
            {cast_real('dr021_weighted_rate')}, {cast_real('dr021_latest_rate')}, {cast_real('dr021_avg_prd')},
            {cast_real('dr1m_weighted_rate')}, {cast_real('dr1m_latest_rate')}, {cast_real('dr1m_avg_prd')},
            {nonblank('fetched_at')}, ?, source_row_num, ?
        FROM stg_sheet_raw_money_market
        WHERE {nonblank('date')} IS NOT NULL
        """,
        ("原始_资金面", migrated_at),
    )

    counts["raw_money_market_from_legacy"] = merge_table(
        conn,
        f"""
        INSERT OR IGNORE INTO raw_money_market (
            date, show_date_cn, source_url,
            dr001_weighted_rate, dr001_latest_rate, dr001_avg_prd,
            dr007_weighted_rate, dr007_latest_rate, dr007_avg_prd,
            dr014_weighted_rate, dr014_latest_rate, dr014_avg_prd,
            dr021_weighted_rate, dr021_latest_rate, dr021_avg_prd,
            dr1m_weighted_rate, dr1m_latest_rate, dr1m_avg_prd,
            fetched_at, source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('date')}, {nonblank('show_date_cn')}, {nonblank('source_url')},
            {cast_real('dr001_weighted_rate')}, {cast_real('dr001_latest_rate')}, {cast_real('dr001_avg_prd')},
            {cast_real('dr007_weighted_rate')}, {cast_real('dr007_latest_rate')}, {cast_real('dr007_avg_prd')},
            {cast_real('dr014_weighted_rate')}, {cast_real('dr014_latest_rate')}, {cast_real('dr014_avg_prd')},
            {cast_real('dr021_weighted_rate')}, {cast_real('dr021_latest_rate')}, {cast_real('dr021_avg_prd')},
            {cast_real('dr1m_weighted_rate')}, {cast_real('dr1m_latest_rate')}, {cast_real('dr1m_avg_prd')},
            {nonblank('fetched_at')}, ?, source_row_num, ?
        FROM stg_sheet_raw_money_legacy
        WHERE {nonblank('date')} IS NOT NULL
        """,
        ("原始_货币", migrated_at),
    )

    counts["raw_futures"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO raw_futures (
            date, t0_last, tf0_last, source, fetched_at,
            source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('date')},
            {cast_real('t0_last')},
            {cast_real('tf0_last')},
            {nonblank('source')},
            {nonblank('fetched_at')},
            ?, source_row_num, ?
        FROM stg_sheet_raw_futures
        WHERE {nonblank('date')} IS NOT NULL
        """,
        ("原始_国债期货", migrated_at),
    )

    counts["raw_life_asset"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO raw_life_asset (
            date, mortgage_rate_est, house_price_tier1, house_price_tier2, house_price_nbs_70city,
            gold_cny, money_fund_7d, deposit_1y, source, fetched_at,
            source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('date')},
            {cast_real('mortgage_rate_est')},
            {cast_real('house_price_tier1')},
            {cast_real('house_price_tier2')},
            {cast_real('house_price_nbs_70city')},
            {cast_real('gold_cny')},
            {cast_real('money_fund_7d')},
            CAST(COALESCE({nonblank('deposit_1y')}, {nonblank('deposit_1y_dup')}) AS REAL),
            COALESCE({nonblank('source')}, {nonblank('source_dup')}),
            COALESCE({nonblank('fetched_at')}, {nonblank('fetched_at_dup')}),
            ?, source_row_num, ?
        FROM stg_sheet_raw_life_asset
        WHERE {nonblank('date')} IS NOT NULL
        """,
        ("原始_民生与资产价格", migrated_at),
    )

    counts["raw_jisilu_etf"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO raw_jisilu_etf (
            snapshot_date, fetched_at, fund_id, fund_nm, index_nm, issuer_nm, price, increase_rt,
            volume_wan, amount_yi, unit_total_yi, discount_rt, fund_nav, nav_dt, estimate_value,
            creation_unit, pe, pb, last_time, last_est_time, is_qdii, is_t0, apply_fee, redeem_fee,
            records_total, source_url, source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('snapshot_date')},
            {nonblank('fetched_at')},
            {nonblank('fund_id')},
            {nonblank('fund_nm')},
            {nonblank('index_nm')},
            {nonblank('issuer_nm')},
            {cast_real('price')},
            {cast_real('increase_rt')},
            {cast_real('volume_wan')},
            {cast_real('amount_yi')},
            {cast_real('unit_total_yi')},
            {cast_real('discount_rt')},
            {cast_real('fund_nav')},
            {nonblank('nav_dt')},
            {cast_real('estimate_value')},
            {cast_real('creation_unit')},
            {cast_real('pe')},
            {cast_real('pb')},
            {nonblank('last_time')},
            {nonblank('last_est_time')},
            {nonblank('is_qdii')},
            {nonblank('is_t0')},
            {cast_real('apply_fee')},
            {cast_real('redeem_fee')},
            {cast_real('records_total')},
            {nonblank('source_url')},
            ?, source_row_num, ?
        FROM stg_sheet_raw_jisilu_etf
        WHERE {nonblank('snapshot_date')} IS NOT NULL
          AND {nonblank('fund_id')} IS NOT NULL
        """,
        ("原始_指数ETF", migrated_at),
    )

    counts["raw_bond_index"] = merge_table(
        conn,
        f"""
        INSERT OR REPLACE INTO raw_bond_index (
            trade_date, index_name, index_code, provider, type_lv1, type_lv2, type_lv3,
            source_url, data_date, dm, y, cons_number, d, v, fetch_status, raw_json, fetched_at, error,
            source_sheet, source_row_num, migrated_at
        )
        SELECT
            {nonblank('trade_date')},
            {nonblank('index_name')},
            {nonblank('index_code')},
            {nonblank('provider')},
            {nonblank('type_lv1')},
            {nonblank('type_lv2')},
            {nonblank('type_lv3')},
            {nonblank('source_url')},
            {nonblank('data_date')},
            {cast_real('dm')},
            {cast_real('y')},
            {cast_int('cons_number')},
            {cast_real('d')},
            {cast_real('v')},
            {nonblank('fetch_status')},
            {nonblank('raw_json')},
            {nonblank('fetched_at')},
            {nonblank('error')},
            ?, source_row_num, ?
        FROM stg_sheet_raw_bond_index
        WHERE {nonblank('trade_date')} IS NOT NULL
          AND {nonblank('index_name')} IS NOT NULL
        """,
        ("原始_债券指数特征", migrated_at),
    )

    return counts


def ensure_ingest_tables(conn: sqlite3.Connection) -> None:
    required = {"ingest_run", "ingest_error"}
    existing = {
        row[0]
        for row in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name IN ('ingest_run','ingest_error')"
        ).fetchall()
    }
    missing = required - existing
    if missing:
        raise RuntimeError(
            f"Missing ingest metadata tables: {sorted(missing)}. Please run init_db.py first."
        )


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    db_path = Path(args.db).expanduser().resolve()

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")
    if not db_path.exists():
        raise FileNotFoundError(f"Database not found: {db_path}. Please run init_db.py first.")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row

    run_id = uuid.uuid4().hex
    started_at = datetime.now().replace(microsecond=0).isoformat(sep=" ")
    rows_read = 0
    rows_written = 0

    try:
        ensure_ingest_tables(conn)
        conn.execute("PRAGMA foreign_keys = ON;")
        conn.execute("BEGIN")

        conn.execute(
            """
            INSERT INTO ingest_run (
                run_id, job_name, source_type, started_at, status, rows_read, rows_written, message
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                "import_from_sheet",
                "xlsx",
                started_at,
                "running",
                0,
                0,
                f"workbook={xlsx_path.name}",
            ),
        )

        for spec in SHEET_SPECS:
            staged = stage_sheet(conn, wb, spec, args.sheet_max_scan_rows)
            rows_read += staged

        merge_counts = merge_staged_data(conn)
        rows_written = sum(merge_counts.values())

        conn.execute(
            """
            UPDATE ingest_run
            SET finished_at = ?, status = ?, rows_read = ?, rows_written = ?, message = ?
            WHERE run_id = ?
            """,
            (
                datetime.now().replace(microsecond=0).isoformat(sep=" "),
                "success",
                rows_read,
                rows_written,
                json.dumps(merge_counts, ensure_ascii=False),
                run_id,
            ),
        )

        conn.commit()

        print("[OK] Sheet import finished.")
        print(f"[OK] Workbook: {xlsx_path}")
        print(f"[OK] Database: {db_path}")
        print(f"[OK] Staged rows: {rows_read}")
        print(f"[OK] Merged rows: {rows_written}")
        for name, count in merge_counts.items():
            print(f"  - {name}: {count}")

    except Exception as exc:
        error_payload = {
            "xlsx": str(xlsx_path),
            "db": str(db_path),
        }
        conn.rollback()
        conn.execute(
            """
            UPDATE ingest_run
            SET finished_at = ?, status = ?, rows_read = ?, rows_written = ?, message = ?
            WHERE run_id = ?
            """,
            (
                datetime.now().replace(microsecond=0).isoformat(sep=" "),
                "failed",
                rows_read,
                rows_written,
                str(exc),
                run_id,
            ),
        )
        conn.execute(
            """
            INSERT INTO ingest_error (
                run_id, source_name, object_name, error_type, error_message, row_payload, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                "xlsx",
                xlsx_path.name,
                exc.__class__.__name__,
                str(exc),
                json.dumps(error_payload, ensure_ascii=False),
                datetime.now().replace(microsecond=0).isoformat(sep=" "),
            ),
        )
        conn.commit()
        raise
    finally:
        conn.close()
        wb.close()


if __name__ == "__main__":
    main()
